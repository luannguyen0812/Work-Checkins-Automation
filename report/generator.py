from datetime import date, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datastore.models import RiskScore, CheckIn
from utils.time_utils import is_us_public_holiday, scheduled_weekdays

COLOR_GREEN  = "00B050"
COLOR_AMBER  = "FFC000"
COLOR_RED    = "FF0000"
COLOR_HEADER = "2F4F8F"
COLOR_GREY   = "D9D9D9"

FILL_GREEN  = PatternFill("solid", fgColor=COLOR_GREEN)
FILL_AMBER  = PatternFill("solid", fgColor=COLOR_AMBER)
FILL_RED    = PatternFill("solid", fgColor=COLOR_RED)
FILL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER)
FILL_GREY   = PatternFill("solid", fgColor=COLOR_GREY)

FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_BOLD   = Font(bold=True)

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr(ws, row, col, value, fill=None, font=None, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.alignment = Alignment(horizontal=align, wrap_text=True)
    cell.border = BORDER
    return cell


def _band_fill(band: str) -> PatternFill:
    return {"GREEN": FILL_GREEN, "AMBER": FILL_AMBER, "RED": FILL_RED}.get(band, FILL_GREY)


def generate_report(iso_week: int, year: int) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_executive_summary(wb, iso_week, year)
    _build_raw_checkins(wb, iso_week, year)
    _build_attendance_rates(wb, iso_week, year)
    _build_trend_lines(wb, iso_week, year)
    _build_heatmap(wb, iso_week, year)
    _build_streaks(wb, iso_week, year)
    _build_risk_scores(wb, iso_week, year)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def report_filename(iso_week: int, year: int, gen_date: date) -> str:
    return f"InternAttendance_W{iso_week:02d}_{gen_date.isoformat()}.xlsx"


def _week_date_range(iso_week: int, year: int) -> tuple[date, date]:
    monday = date.fromisocalendar(year, iso_week, 1)
    sunday = monday + timedelta(days=6)
    return monday, sunday


def _load_data(iso_week: int, year: int):
    from datastore import sheets
    from datastore.queries import compute_all_risk_scores
    _, week_end = _week_date_range(iso_week, year)
    checkins = sheets.get_checkins_for_week(iso_week, year)
    scores = compute_all_risk_scores(iso_week, year, as_of=week_end)
    return checkins, scores


def _build_executive_summary(wb, iso_week: int, year: int) -> None:
    from utils.time_utils import edt_now
    ws = wb.create_sheet("Executive Summary")
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    monday, sunday = _week_date_range(iso_week, year)
    checkins, scores = _load_data(iso_week, year)

    total = len(scores)
    avg_rate = round(sum(s.war for s in scores) / total * 100, 1) if scores else 0.0
    green = [s for s in scores if s.war >= 0.85]
    amber = [s for s in scores if 0.70 <= s.war < 0.85]
    red = [s for s in scores if s.war < 0.70]

    # Title block
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = f"Intern Attendance Report — Week {iso_week} ({monday.strftime('%b %d')} – {sunday.strftime('%b %d, %Y')})"
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = FILL_HEADER
    title.alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 30

    # KPI cards
    kpis = [
        ("Total Interns", total, None),
        ("Avg Attendance", f"{avg_rate}%", None),
        ("GREEN", len(green), FILL_GREEN),
        ("RED", len(red), FILL_RED),
    ]
    for col, (label, val, fill) in enumerate(kpis, start=1):
        ws.cell(row=2, column=col, value=label).font = FONT_BOLD
        cell = ws.cell(row=3, column=col, value=val)
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal="center")
        if fill:
            cell.fill = fill
            if fill in (FILL_GREEN, FILL_RED):
                cell.font = Font(bold=True, size=16, color="FFFFFF")

    ws.row_dimensions[3].height = 30

    # Weekly attendance watchlist. Longer-term composite risk stays in Risk Scores.
    ws.cell(row=5, column=1, value="Weekly Attendance Watchlist").font = Font(bold=True, size=12)
    headers = ["Name", "Attendance", "Streak", "Attendance Band", "Action"]
    for col, h in enumerate(headers, start=1):
        _hdr(ws, 6, col, h, fill=FILL_HEADER, font=FONT_HEADER)

    watchlist = sorted(red + amber, key=lambda s: (s.war, s.full_name))[:5]
    for row_offset, s in enumerate(watchlist, start=7):
        band = "GREEN" if s.war >= 0.85 else ("AMBER" if s.war >= 0.70 else "RED")
        action = "Follow up" if band == "RED" else "Monitor"
        vals = [s.full_name, f"{s.war*100:.0f}%", str(s.cas), band, action]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_offset, column=col, value=v)
            cell.border = BORDER
            if col == 4:
                cell.fill = _band_fill(band)
                if band != "AMBER":
                    cell.font = Font(color="FFFFFF")

    # Narrative
    narrative_row = 14
    ws.cell(row=narrative_row, column=1, value="Executive Summary").font = Font(bold=True, size=12)
    narrative = _get_narrative(iso_week, year, scores)
    ws.merge_cells(f"A{narrative_row+1}:D{narrative_row+5}")
    narr_cell = ws[f"A{narrative_row+1}"]
    narr_cell.value = narrative
    narr_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Footer
    footer_row = narrative_row + 7
    ws.cell(row=footer_row, column=1, value=f"Generated: {edt_now().strftime('%Y-%m-%d %H:%M')} EDT").font = Font(italic=True, size=9, color="888888")


def _get_narrative(iso_week: int, year: int, scores: list[RiskScore]) -> str:
    import os
    monday, sunday = _week_date_range(iso_week, year)
    total = len(scores)
    avg_rate = round(sum(s.war for s in scores) / total * 100, 1) if scores else 0.0
    red = [s for s in scores if s.war < 0.70]
    amber = [s for s in scores if 0.70 <= s.war < 0.85]
    green = [s for s in scores if s.war >= 0.85]
    red_table = "\n".join(f"  {s.full_name}: {s.war*100:.0f}% attendance, {s.cas} day streak" for s in red[:5])
    trend_notes = f"{len(red)} interns below 70% weekly attendance; {len(amber)} between 70% and 85%"

    if os.environ.get("ANTHROPIC_API_KEY", "").strip():
        try:
            from report.summary import generate_narrative
            return generate_narrative(
                week_number=iso_week,
                date_range=f"{monday.strftime('%b %d')} – {sunday.strftime('%b %d, %Y')}",
                total_interns=total,
                avg_rate=avg_rate,
                green_count=len(green),
                amber_count=len(amber),
                red_count=len(red),
                red_interns_table=red_table or "None",
                trend_notes=trend_notes,
            )
        except Exception:
            pass

    # Fallback plain narrative
    lines = [
        f"Week {iso_week} summary ({monday.strftime('%b %d')} – {sunday.strftime('%b %d, %Y')}).",
        f"Cohort average attendance: {avg_rate}%. {total} interns tracked.",
        f"Weekly attendance: {len(green)} GREEN, {len(amber)} AMBER, {len(red)} RED.",
    ]
    if red:
        names = ", ".join(s.full_name for s in red[:3])
        lines.append(f"RED attendance interns requiring follow-up: {names}{'...' if len(red) > 3 else ''}.")
    if amber:
        lines.append(f"{len(amber)} AMBER interns should be monitored closely next week.")
    return " ".join(lines)


def _build_raw_checkins(wb, iso_week: int, year: int) -> None:
    from datastore import sheets
    ws = wb.create_sheet("Raw Check-Ins")
    checkins = sheets.get_checkins_for_week(iso_week, year)

    headers = ["Date", "Intern ID", "Telegram ID", "Full Name", "UTC Timestamp", "EDT Timestamp",
               "Message", "Message ID", "Validated", "Late", "Week"]
    for col, h in enumerate(headers, start=1):
        _hdr(ws, 1, col, h, fill=FILL_HEADER, font=FONT_HEADER)
        ws.column_dimensions[get_column_letter(col)].width = 18

    for row, c in enumerate(checkins, start=2):
        vals = [
            c.date.isoformat(), c.intern_id, c.telegram_user_id, c.full_name,
            str(c.checkin_timestamp_utc), str(c.checkin_timestamp_edt),
            c.message_text, c.message_id, c.validated, c.late, c.week_number,
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = BORDER
            if col == 10 and v is True:  # late
                cell.fill = FILL_AMBER
            elif col == 10 and v is False:
                cell.fill = FILL_GREEN

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _working_days(iso_week: int, year: int) -> list[date]:
    """All 7 calendar days of the ISO week. Which of these count for a given
    intern depends on their own schedule — see scheduled_weekdays()."""
    monday = date.fromisocalendar(year, iso_week, 1)
    return [monday + timedelta(days=i) for i in range(7)]


def _scheduled_days_for_intern(intern, days: list[date]) -> list[date]:
    allowed = scheduled_weekdays(intern)
    return [
        d
        for d in days
        if d.weekday() in allowed
        and intern.start_date <= d <= intern.end_date
        and not is_us_public_holiday(d)
    ]


def _attendance_days_for_intern(intern, days: list[date], checked: set[date]) -> list[date]:
    scheduled = set(_scheduled_days_for_intern(intern, days))
    valid_checked = {
        d for d in checked
        if d in days and intern.start_date <= d <= intern.end_date
    }
    return sorted(scheduled | valid_checked)


def _build_attendance_rates(wb, iso_week: int, year: int) -> None:
    from datastore import sheets
    from report.charts import add_attendance_bar_chart
    ws = wb.create_sheet("Attendance Rates")

    checkins = sheets.get_checkins_for_week(iso_week, year)
    interns = [i for i in sheets.get_all_interns() if i.active]
    days = _working_days(iso_week, year)
    day_labels = [d.strftime("%a %m/%d") for d in days]

    headers = ["Intern"] + day_labels + ["Rate %", "Risk"]
    for col, h in enumerate(headers, start=1):
        _hdr(ws, 1, col, h, fill=FILL_HEADER, font=FONT_HEADER)
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 22

    checked_map: dict[str, set[date]] = {}
    for c in checkins:
        if c.validated:
            checked_map.setdefault(c.intern_id, set()).add(c.date)

    rows_written = 0
    for row_offset, intern in enumerate(sorted(interns, key=lambda i: i.full_name), start=2):
        checked = checked_map.get(intern.intern_id, set())
        intern_days = _attendance_days_for_intern(intern, days, checked)
        intern_day_set = set(intern_days)
        rate = len(checked & set(intern_days)) / len(intern_days) * 100 if intern_days else 0
        band = "GREEN" if rate >= 85 else ("AMBER" if rate >= 70 else "RED")

        ws.cell(row=row_offset, column=1, value=intern.full_name).border = BORDER
        for col_offset, d in enumerate(days, start=2):
            if d not in intern_day_set:
                cell = ws.cell(row=row_offset, column=col_offset, value="—")
                cell.fill = FILL_GREY
            else:
                cell = ws.cell(row=row_offset, column=col_offset, value="✅" if d in checked else "❌")
                cell.fill = FILL_GREEN if d in checked else FILL_RED
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

        rate_cell = ws.cell(row=row_offset, column=len(days) + 2, value=round(rate, 1))
        rate_cell.fill = _band_fill(band)
        rate_cell.border = BORDER
        if band != "AMBER":
            rate_cell.font = Font(color="FFFFFF")

        band_cell = ws.cell(row=row_offset, column=len(days) + 3, value=band)
        band_cell.fill = _band_fill(band)
        band_cell.border = BORDER
        if band != "AMBER":
            band_cell.font = Font(color="FFFFFF")
        rows_written += 1

    if rows_written:
        add_attendance_bar_chart(ws, name_col=1, rate_col=len(days) + 2, data_rows=rows_written)


def _build_trend_lines(wb, iso_week: int, year: int) -> None:
    from datastore import sheets
    ws = wb.create_sheet("Trend Lines")
    interns = [i for i in sheets.get_all_interns() if i.active]

    # Build 4-week rolling data
    now = date.fromisocalendar(year, iso_week, 5)  # Friday of the report week
    week_refs = []
    for w in range(3, -1, -1):
        ref = now - timedelta(weeks=w)
        ref_iso = ref.isocalendar()
        week_refs.append((ref_iso.week, ref_iso.year))

    week_labels = [f"W{w}" for w, y in week_refs]
    headers = ["Intern"] + week_labels
    for col, h in enumerate(headers, start=1):
        _hdr(ws, 1, col, h, fill=FILL_HEADER, font=FONT_HEADER)
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 22

    # Pre-fetch all 4 weeks
    week_data: dict[tuple, list[CheckIn]] = {}
    for w, y in week_refs:
        week_data[(w, y)] = sheets.get_checkins_for_week(w, y)

    rows_written = 0
    for row_offset, intern in enumerate(sorted(interns, key=lambda i: i.full_name), start=2):
        ws.cell(row=row_offset, column=1, value=intern.full_name).border = BORDER
        for col_offset, (w, y) in enumerate(week_refs, start=2):
            checked = {c.date for c in week_data[(w, y)] if c.intern_id == intern.intern_id and c.validated}
            wd = _attendance_days_for_intern(intern, _working_days(w, y), checked)
            rate = round(len({d for d in wd if d in checked}) / len(wd) * 100, 1) if wd else 0
            cell = ws.cell(row=row_offset, column=col_offset, value=rate)
            cell.border = BORDER
        rows_written += 1


def _build_heatmap(wb, iso_week: int, year: int) -> None:
    from datastore import sheets
    from utils.time_utils import edt_now
    ws = wb.create_sheet("Heatmap")

    checkins = sheets.get_checkins_for_week(iso_week, year)
    interns = [i for i in sheets.get_all_interns() if i.active]
    days = _working_days(iso_week, year)
    _, week_end = _week_date_range(iso_week, year)
    today = min(edt_now().date(), week_end)

    headers = ["Intern"] + [d.strftime("%a %m/%d") for d in days]
    for col, h in enumerate(headers, start=1):
        _hdr(ws, 1, col, h, fill=FILL_HEADER, font=FONT_HEADER)
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 22

    late_map: dict[tuple, bool] = {(c.intern_id, c.date): c.late for c in checkins}
    checked_map: dict[str, set[date]] = {}
    for c in checkins:
        if c.validated:
            checked_map.setdefault(c.intern_id, set()).add(c.date)

    sorted_interns = sorted(interns, key=lambda i: len(checked_map.get(i.intern_id, set())))
    for row_offset, intern in enumerate(sorted_interns, start=2):
        ws.cell(row=row_offset, column=1, value=intern.full_name).border = BORDER
        checked = checked_map.get(intern.intern_id, set())
        intern_day_set = set(_attendance_days_for_intern(intern, days, checked))
        for col_offset, d in enumerate(days, start=2):
            if d not in intern_day_set:
                symbol, fill = "—", FILL_GREY
            elif d > today:
                symbol, fill = "—", FILL_GREY
            elif d in checked:
                late = late_map.get((intern.intern_id, d), False)
                symbol, fill = ("⏰", FILL_AMBER) if late else ("✅", FILL_GREEN)
            else:
                symbol, fill = "❌", FILL_RED
            cell = ws.cell(row=row_offset, column=col_offset, value=symbol)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER


def _build_streaks(wb, iso_week: int, year: int) -> None:
    from datastore import sheets
    from datastore.queries import compute_all_risk_scores
    from utils.time_utils import edt_now
    ws = wb.create_sheet("Streaks & Gaps")

    interns = [i for i in sheets.get_all_interns() if i.active]
    _, week_end = _week_date_range(iso_week, year)
    scores_map = {s.intern_id: s for s in compute_all_risk_scores(iso_week, year, as_of=week_end)}
    checkins = sheets.get_checkins_for_week(iso_week, year)
    days = _working_days(iso_week, year)

    headers = ["Intern", "Days Present", "Days Absent", "Current Absence Streak"] + [d.strftime("%a %m/%d") for d in days]
    for col, h in enumerate(headers, start=1):
        _hdr(ws, 1, col, h, fill=FILL_HEADER, font=FONT_HEADER)
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["A"].width = 22

    checked_map: dict[str, set[date]] = {}
    for c in checkins:
        if c.validated:
            checked_map.setdefault(c.intern_id, set()).add(c.date)

    for row_offset, intern in enumerate(sorted(interns, key=lambda i: i.full_name), start=2):
        checked = checked_map.get(intern.intern_id, set())
        intern_days = _attendance_days_for_intern(intern, days, checked)
        intern_day_set = set(intern_days)
        present = len({d for d in intern_days if d in checked})
        absent = len(intern_days) - present
        cas = scores_map.get(intern.intern_id).cas if intern.intern_id in scores_map else 0

        row_vals = [intern.full_name, present, absent, cas]
        for col, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_offset, column=col, value=v)
            cell.border = BORDER
            if col == 4 and v >= 3:
                cell.fill = FILL_RED
                cell.font = Font(bold=True, color="FFFFFF")

        today = min(edt_now().date(), week_end)
        for col_offset, d in enumerate(days, start=5):
            if d not in intern_day_set:
                val, fill = "—", FILL_GREY
            elif d > today:
                val, fill = "—", FILL_GREY
            elif d in checked:
                val, fill = "✅", FILL_GREEN
            else:
                val, fill = "❌", FILL_RED
            cell = ws.cell(row=row_offset, column=col_offset, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER


def _build_risk_scores(wb, iso_week: int, year: int) -> None:
    from datastore.queries import compute_all_risk_scores
    from report.charts import add_risk_scatter_chart
    ws = wb.create_sheet("Risk Scores")

    _, week_end = _week_date_range(iso_week, year)
    scores = sorted(compute_all_risk_scores(iso_week, year, as_of=week_end), key=lambda s: s.risk_score)

    headers = ["Intern", "WAR %", "RAR %", "CAS", "LCR %", "Risk Score", "Band"]
    for col, h in enumerate(headers, start=1):
        _hdr(ws, 1, col, h, fill=FILL_HEADER, font=FONT_HEADER)
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["A"].width = 22

    for row_offset, s in enumerate(scores, start=2):
        vals = [
            s.full_name,
            round(s.war * 100, 1),
            round(s.rar * 100, 1),
            s.cas,
            round(s.lcr * 100, 1),
            round(s.risk_score, 3),
            s.risk_band,
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_offset, column=col, value=v)
            cell.border = BORDER
            if col == 7:
                cell.fill = _band_fill(s.risk_band)
                if s.risk_band != "AMBER":
                    cell.font = Font(color="FFFFFF")

    if scores:
        add_risk_scatter_chart(ws, war_col=2, rar_col=3, data_rows=len(scores))

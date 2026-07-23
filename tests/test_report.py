from datetime import date, datetime
from io import BytesIO

import openpyxl

from datastore.models import CheckIn, Intern, RiskScore
from report.generator import generate_report
from utils.time_utils import is_us_public_holiday


def _intern(intern_id: str, name: str, schedule_json: str | None = None) -> Intern:
    return Intern(
        intern_id=intern_id,
        full_name=name,
        telegram_user_id=123,
        cohort="2026-Summer",
        start_date=date(2026, 6, 1),
        end_date=date(2026, 8, 31),
        schedule_json=schedule_json,
    )


def _checkin(intern: Intern, day: date) -> CheckIn:
    ts = datetime(day.year, day.month, day.day, 13, 0)
    return CheckIn(
        date=day,
        intern_id=intern.intern_id,
        telegram_user_id=intern.telegram_user_id,
        full_name=intern.full_name,
        checkin_timestamp_utc=ts,
        checkin_timestamp_edt=ts,
        message_text="I'm online",
        message_id=1,
        validated=True,
        late=False,
        week_number=day.isocalendar().week,
    )


def _score(intern: Intern) -> RiskScore:
    return RiskScore(
        intern_id=intern.intern_id,
        full_name=intern.full_name,
        war=1.0,
        rar=1.0,
        cas=0,
        lcr=0.0,
        risk_score=1.0,
        risk_band="GREEN",
    )


def test_weekend_workers_are_scored_on_saturday_and_sunday(monkeypatch):
    weekday_only = _intern("weekday", "Weekday Only")
    weekend_worker = _intern(
        "weekend",
        "Weekend Worker",
        '[{"days":["Sat","Sun"],"start":"10:00","end":"18:00"}]',
    )
    saturday = date(2026, 6, 27)
    sunday = date(2026, 6, 28)
    checkins = [_checkin(weekend_worker, saturday), _checkin(weekend_worker, sunday)]

    from datastore import sheets
    import datastore.queries

    monkeypatch.setattr(sheets, "get_all_interns", lambda: [weekday_only, weekend_worker])
    monkeypatch.setattr(sheets, "get_checkins_for_week", lambda week, year: checkins)
    monkeypatch.setattr(
        datastore.queries,
        "compute_all_risk_scores",
        lambda week, year, as_of=None: [_score(weekday_only), _score(weekend_worker)],
    )

    wb = openpyxl.load_workbook(BytesIO(generate_report(26, 2026)))
    rates = wb["Attendance Rates"]
    headers = [rates.cell(1, col).value for col in range(1, rates.max_column + 1)]
    assert "Sat 06/27" in headers
    assert "Sun 06/28" in headers

    rows = {
        rates.cell(row, 1).value: {
            headers[col - 1]: rates.cell(row, col).value
            for col in range(2, rates.max_column + 1)
        }
        for row in range(2, rates.max_row + 1)
    }

    assert rows["Weekend Worker"]["Sat 06/27"] == "✅"
    assert rows["Weekend Worker"]["Sun 06/28"] == "✅"
    assert rows["Weekend Worker"]["Rate %"] == 100
    assert rows["Weekday Only"]["Sat 06/27"] == "—"
    assert rows["Weekday Only"]["Sun 06/28"] == "—"


def test_off_schedule_checkin_counts_as_present(monkeypatch):
    intern = _intern(
        "off_schedule",
        "Off Schedule",
        '[{"days":["Wed"],"start":"10:00","end":"18:00"}]',
    )
    monday = date(2026, 6, 22)
    checkins = [_checkin(intern, monday)]

    from datastore import sheets
    import datastore.queries

    monkeypatch.setattr(sheets, "get_all_interns", lambda: [intern])
    monkeypatch.setattr(sheets, "get_checkins_for_week", lambda week, year: checkins)
    monkeypatch.setattr(
        datastore.queries,
        "compute_all_risk_scores",
        lambda week, year, as_of=None: [_score(intern)],
    )

    wb = openpyxl.load_workbook(BytesIO(generate_report(26, 2026)))
    rates = wb["Attendance Rates"]
    headers = [rates.cell(1, col).value for col in range(1, rates.max_column + 1)]
    row = {
        headers[col - 1]: rates.cell(2, col).value
        for col in range(2, rates.max_column + 1)
    }

    assert row["Mon 06/22"] == "✅"
    assert row["Wed 06/24"] == "❌"
    assert row["Rate %"] == 50


def test_actual_and_observed_fixed_holidays_are_public_holidays():
    assert is_us_public_holiday(date(2026, 7, 3))
    assert is_us_public_holiday(date(2026, 7, 4))
